"""
rdg/core/exporter.py

Exports the generated dataset to:
  1. Formatted Excel workbook (.xlsx)  — multiple sheets
  2. Raw CSV
  3. SPSS-ready CSV (numeric codes only)
  4. Plain-text validation report

The exporter is study-agnostic. Column names are discovered from the data.
Study-specific SPSS encoding maps are passed in as arguments.
"""
from __future__ import annotations
import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ───────────────────────────────────────────
_NAVY   = "1F3864"
_TEAL   = "17375E"
_GOLD   = "FFC000"
_LTBLUE = "DCE6F1"
_LTGREY = "F2F2F2"
_WHITE  = "FFFFFF"
_GREEN  = "375623"
_ORANGE = "974706"
_RED    = "C00000"


def _style_header(ws, row_num: int, bg: str = _NAVY, fg: str = _WHITE) -> None:
    fill   = PatternFill("solid", fgColor=bg)
    font   = Font(bold=True, color=fg, size=10)
    border = Border(
        bottom=Side(style="medium", color=_GOLD),
        top=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"),
    )
    for cell in ws[row_num]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border


def _alt_rows(ws, start: int, end: int, n_cols: int) -> None:
    f_even = PatternFill("solid", fgColor=_LTBLUE)
    f_odd  = PatternFill("solid", fgColor=_LTGREY)
    for r in range(start, end + 1):
        fill = f_even if r % 2 == 0 else f_odd
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).fill = fill


def _autofit(ws, min_w: int = 8, max_w: int = 42) -> None:
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = (
            min(max(width + 2, min_w), max_w)
        )


def _section_header(ws, title: str, n_cols: int = 4) -> None:
    ws.append([title] + [""] * (n_cols - 1))
    row = ws.max_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1)
    cell.fill      = PatternFill("solid", fgColor=_NAVY)
    cell.font      = Font(bold=True, color=_WHITE, size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")


# ─────────────────────────────────────────────────────────────

def export(
    demographics:       list[dict],
    questionnaire_rows: list[dict],
    observations:       list[dict],
    validation_report:  dict,
    output_dir:         str | Path,
    study_title:        str = "Research Study",
    seed:               int = 42,
    spss_maps:          dict[str, dict[str, int]] | None = None,
    codebook_rows:      list[tuple] | None = None,
) -> dict[str, Path]:
    """
    Export all outputs and return a dict of {label: Path}.
    """
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    q_map  = {r["respondent_id"]: r for r in questionnaire_rows}
    ob_map = {r["respondent_id"]: r for r in observations}
    merged = [{**d, **q_map[d["respondent_id"]], **ob_map[d["respondent_id"]]}
              for d in demographics]

    # ── Excel ────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _sheet_raw(wb, merged)
    _sheet_questionnaire(wb, demographics, questionnaire_rows)
    _sheet_observation(wb, demographics, observations)
    _sheet_summary(wb, demographics, questionnaire_rows)
    _sheet_frequency(wb, demographics, questionnaire_rows)
    if codebook_rows:
        _sheet_codebook(wb, codebook_rows)
    _sheet_validation(wb, validation_report)

    xl_path  = out / f"dataset_{ts}.xlsx"
    csv_path = out / f"raw_{ts}.csv"
    spss_path = out / f"spss_{ts}.csv"
    rpt_path  = out / f"validation_{ts}.txt"

    wb.save(xl_path)

    # ── Raw CSV ──────────────────────────────────────────────
    if merged:
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=merged[0].keys())
            w.writeheader(); w.writerows(merged)

    # ── SPSS CSV ─────────────────────────────────────────────
    if spss_maps:
        _export_spss(merged, spss_path, spss_maps)
    else:
        spss_path.write_text("# No SPSS encoding map provided.\n")

    # ── Validation text ──────────────────────────────────────
    _write_validation_text(validation_report, rpt_path, seed, study_title)

    return {"excel": xl_path, "csv": csv_path, "spss": spss_path, "report": rpt_path}


# ─────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────

def _get_col_groups(row: dict) -> tuple[list, list, list, list, list]:
    demo  = [k for k in row if not (k.startswith("S") and "Q" in k)
             and not k.startswith("mean_") and k != "overall_mean"
             and k != "satisfaction_category"
             and k not in ("facility_id","obs_yes_count")
             and not any(row[k] == "Yes" or row[k] == "No" for _ in [0])]
    likert = [k for k in row if k.startswith("S") and "Q" in k]
    means  = [k for k in row if k.startswith("mean_") or k == "overall_mean"]
    sat    = ["satisfaction_category"] if "satisfaction_category" in row else []
    obs    = [k for k in row if k not in demo + likert + means + sat
              and k not in ("respondent_id",)]
    return demo, likert, means, sat, obs


def _sheet_raw(wb, merged):
    ws = wb.create_sheet("Raw Dataset")
    ws.freeze_panes = "A2"
    if not merged:
        return
    headers = list(merged[0].keys())
    ws.append(headers); _style_header(ws, 1)
    for row in merged:
        ws.append([row.get(h, "") for h in headers])
    _alt_rows(ws, 2, len(merged) + 1, len(headers))
    _autofit(ws)


def _sheet_questionnaire(wb, demographics, questionnaire_rows):
    ws = wb.create_sheet("Questionnaire Data")
    ws.freeze_panes = "A2"
    q_cols  = [k for k in questionnaire_rows[0] if k.startswith("S") and "Q" in k]
    means   = [k for k in questionnaire_rows[0] if k.startswith("mean_") or k == "overall_mean"]
    headers = ["respondent_id"] + q_cols + means + ["satisfaction_category"]
    ws.append(headers); _style_header(ws, 1)
    q_map = {r["respondent_id"]: r for r in questionnaire_rows}
    for d in demographics:
        q = q_map[d["respondent_id"]]
        ws.append([d["respondent_id"]] + [q.get(c) for c in q_cols + means] + [q.get("satisfaction_category")])
    _alt_rows(ws, 2, len(demographics) + 1, len(headers))
    _autofit(ws)


def _sheet_observation(wb, demographics, observations):
    ws = wb.create_sheet("Observation Checklist")
    ws.freeze_panes = "A2"
    obs_keys = [k for k in observations[0] if k not in ("respondent_id","facility_id")]
    headers  = ["respondent_id","facility_id"] + obs_keys
    ws.append(headers); _style_header(ws, 1)
    ob_map = {r["respondent_id"]: r for r in observations}
    for d in demographics:
        ob = ob_map[d["respondent_id"]]
        ws.append([d["respondent_id"], ob.get("facility_id","")] + [ob.get(k,"") for k in obs_keys])
    _alt_rows(ws, 2, len(demographics) + 1, len(headers))
    _autofit(ws)


def _sheet_summary(wb, demographics, questionnaire_rows):
    ws = wb.create_sheet("Summary Statistics")
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 8

    def block(title, rows):
        _section_header(ws, title)
        ws.append(["Variable","Mean","SD","N"]); _style_header(ws, ws.max_row, _TEAL, _WHITE)
        for r in rows:
            ws.append(r)
        ws.append([])

    numeric_demo = {k: [d[k] for d in demographics if isinstance(d.get(k), (int, float))]
                    for k in demographics[0]}
    num_rows = [(k, round(np.mean(v),2), round(np.std(v),2), len(v))
                for k, v in numeric_demo.items() if v]
    if num_rows:
        block("DEMOGRAPHIC STATISTICS", num_rows)

    section_keys = sorted({k.split("_")[1] for k in questionnaire_rows[0] if k.startswith("mean_")})
    sec_rows = []
    for s in section_keys:
        vals = [q[f"mean_{s}"] for q in questionnaire_rows]
        sec_rows.append([f"Section {s} mean", round(np.mean(vals),3), round(np.std(vals),3), len(vals)])
    overall = [q["overall_mean"] for q in questionnaire_rows]
    sec_rows.append(["Overall Satisfaction Mean", round(np.mean(overall),3), round(np.std(overall),3), len(overall)])
    block("SATISFACTION SCORE STATISTICS  (1=Very Dissatisfied → 5=Very Satisfied)", sec_rows)
    _autofit(ws)


def _sheet_frequency(wb, demographics, questionnaire_rows):
    from collections import Counter
    ws = wb.create_sheet("Frequency Tables")

    def freq(title, values):
        ws.append([title, "Frequency", "Percent"])
        _style_header(ws, ws.max_row, _TEAL, _WHITE)
        c = Counter(values); total = len(values)
        for k, v in sorted(c.items(), key=lambda x: -x[1]):
            ws.append([k, v, round(v/total*100, 1)])
        ws.append(["Total", total, 100.0]); ws.append([])

    categorical = [k for k in demographics[0]
                   if isinstance(demographics[0][k], str) and k != "respondent_id"]
    for field in categorical:
        freq(field.replace("_"," ").title(), [d[field] for d in demographics])
    freq("Satisfaction Category", [q["satisfaction_category"] for q in questionnaire_rows])
    _autofit(ws)


def _sheet_codebook(wb, codebook_rows):
    ws = wb.create_sheet("Codebook")
    headers = ["Variable","Label","Type","Scale","Values / Range","Notes"]
    ws.append(headers); _style_header(ws, 1)
    for row in codebook_rows:
        ws.append(list(row))
    _autofit(ws)


def _sheet_validation(wb, report):
    ws = wb.create_sheet("Validation Report")
    ws.append(["Status","Check"]); _style_header(ws, 1)
    colour_map = {"passed": _GREEN, "warnings": _ORANGE, "errors": _RED}
    labels     = {"passed": "✓ PASS", "warnings": "⚠ WARN", "errors": "✗ ERR"}
    for key in ("passed", "warnings", "errors"):
        for item in report.get(key, []):
            r = ws.max_row + 1
            ws.append([labels[key], item])
            ws.cell(row=r, column=1).font = Font(bold=True, color=colour_map[key])
    _autofit(ws)


# ─────────────────────────────────────────────────────────────
# SPSS numeric export
# ─────────────────────────────────────────────────────────────

def _export_spss(merged: list[dict], path: Path, spss_maps: dict[str, dict]) -> None:
    if not merged:
        return
    def encode(row: dict) -> dict:
        out: dict = {}
        for k, v in row.items():
            if k in spss_maps:
                out[k.upper()[:8]] = spss_maps[k].get(str(v), 9)
            elif isinstance(v, (int, float)):
                out[k.upper()[:8]] = v
            elif isinstance(v, str) and (v == "Yes" or v == "No"):
                out[k.upper()[:8]] = 1 if v == "Yes" else 0
        return out
    encoded = [encode(r) for r in merged]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=encoded[0].keys())
        w.writeheader(); w.writerows(encoded)


# ─────────────────────────────────────────────────────────────

def _write_validation_text(report, path, seed, study_title):
    lines = [
        "=" * 60,
        "RESEARCH DATASET GENERATOR — VALIDATION REPORT",
        f"Study   : {study_title}",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Seed    : {seed}",
        "=" * 60, "",
        f"PASSED  : {report['summary']['n_passed']}",
        f"WARNINGS: {report['summary']['n_warnings']}",
        f"ERRORS  : {report['summary']['n_errors']}",
        f"READY   : {'YES' if report['summary']['ready_to_export'] else 'NO'}",
        "",
        "── PASSED ──────────────────────────────────────────",
    ]
    for item in report.get("passed", []):
        lines.append(f"  ✓  {item}")
    lines += ["", "── WARNINGS ─────────────────────────────────────────"]
    for item in report.get("warnings", []):
        lines.append(f"  ⚠  {item}")
    lines += ["", "── ERRORS ───────────────────────────────────────────"]
    for item in report.get("errors", []):
        lines.append(f"  ✗  {item}")
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")
